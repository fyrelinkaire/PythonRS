#!/usr/bin/env python
"""
Created on Tue Sep 27 09:06:32 2022
Last Maintained on Wed Feb 22 12:51:50 2023
@author: ALiu3
"""

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import colors
from openpyxl.styles import Font, Alignment, PatternFill #, Color, Border
#from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import FormulaRule #, ColorScaleRule, CellIsRule, Rule
from pathlib import Path
import pandas as pd
import os
import time
import re


'''
-----------------------------------------------------------------------------
                            Initialization
-----------------------------------------------------------------------------
'''
p = Path('.')                                                                  #Pathlib module
wb = Workbook()                                                                #Open Patient QA worksheet
ws = wb.active                                                                 #Set active worksheet

'''
-----------------------------------------------------------------------------
                            Sheet 1 Formatting
-----------------------------------------------------------------------------
'''
colA = 1  #A
colB = 2  #B
colC = 3  #C
colD = 4  #D
colE = 5  #E
colF = 6  #f
colG = 7  #G
colH = 8  #H
colI = 9  #I
ws['A4'] = 'Patient Name'
ws['A5'] = 'Patient URN'
ws['A6'] = 'Plan Name'
ws['A7'] = 'Plan folder'
ws['A8'] = 'Machine'
ws['A10'] = 'Phantom Used'
ws['A11'] = 'Dose Calibration'
ws['A12'] = 'Array Calibration'
ws['A13'] = 'Dose Rate (MU/min)'
ws['G10'] = 'Threshold %'
ws['G11'] = '% Dose Difference'
ws['G12'] = 'Dist to Agreement'
ws['G13'] = '3D DTA Mode'
ws['C15'] = 'Field'
ws['D15'] = 'Dose' + ' ' u'\u2206' + '%'
ws['E15'] = 'Passing %'
ws['F15'] = '#pts y>2'
ws['G15'] = 'Global'
ws['H15'] = 'P/F'
ws['A27'] = '(1) Absolute point dose difference is ≤ 3%'
ws['A28'] = '(2) Absolute point dose difference is ≤ 5%'
ws['A29'] = '(3) % points passing the absolute y test is ≥ 90%'
ws['A30'] = '(4) No points with y > 2 on the absolute y test'
ws['A32'] = 'Insert comments here…'
ws['I27'] = 'Passes if (2) & (3) or (4) are satisfied if global off'
ws['I28'] = 'Passes if (1) & (3) & (4) are satisfied if global on'
ws['I29'] = 'Otherwise fails'

title = ws['A2'] # title 
title.font = Font(sz=18)
ws['A2'] = 'Patient QA Worksheet'

def colorCells(min_row, max_row, min_col, max_col):
    
    for rows in ws.iter_rows( min_row= min_row, 
                              max_row= max_row,
                              min_col= min_col,
                              max_col= max_col):
        for cell in rows:
            cell.fill = PatternFill(patternType='solid', fgColor=colors.Color(rgb='EFF7FF'))
    return cell

colorCells( 4,  8, colC, colH)                                                 #color cells C4:C8 
colorCells(10, 13, colC, colC)                                                 #color cells C10:C13 
colorCells(10, 13, colH, colH)                                                 #color cells H10:C13 
colorCells(16, 25, colC, colH)                                                 #color cells C16:C25
colorCells(32, 34, colA, colI)                                                 #color cells A32:I34         

def mergeCells(min_row, max_row, min_col, max_col):

    for rows in ws.iter_rows(min_row= min_row, 
                             max_row= max_row,
                             min_col= min_col,
                             max_col= max_col):
            for cell in range(min_row, max_row + 1):
                ws.merge_cells(start_row = cell,
                start_column = min_col,
                end_row = cell,
                end_column = max_col,
                )   
    return cell

mergeCells( 4,  8, colA, colB)                                                 #replaces A4:A8, merges col1 + col2
mergeCells(10, 13, colA, colB)                                                 #replaces A10:A13, merges col1 + col2
mergeCells( 4,  8, colC, colH)                                                 #replaces C4:C8, merges col4 + col8
mergeCells(10, 13, colC, colD)                                                 #replaces C10:C13, merges col4 + col8
mergeCells( 2,  2, colA, colI)                                                 #replaces A2:I2, merges col1 + col 9
def alignCells(min_row, max_row, column, value):

    for rows in ws.iter_rows(min_row= min_row, 
                             max_row= max_row,
                             min_col= column,
                             ):
            for cell in range(min_row, max_row + 1):
                ws.cell(row= cell, column= column).alignment = Alignment(horizontal= value)
                
    return cell

alignCells( 1,  8,  colA, 'right')                                             #right aligns A4:A8
alignCells(10, 13,  colA, 'right')                                             #right aligns A10:A13
alignCells(10, 13,  colG, 'right')                                             #right aligns G10:G13
alignCells(27, 29,  colI, 'right')                                             #right aligns I10:I13
alignCells(10, 13,  colC, 'center')                                            #center aligns C10:C13
alignCells(10, 13,  colH, 'center')                                            #center aligns H10:H13
alignCells( 4,  8,  colC, 'left')                                              #center aligns C4:C8
alignCells( 2,  2,  colA, 'center')                                            #center aligns A2:I2

def number_formatCells(min_row, max_row, column):

    for rows in ws.iter_rows(min_row= min_row, 
                             max_row= max_row,
                             min_col= column,
                             ):
            for cell in range(min_row, max_row + 1):
                ws.cell(row= cell, column= column).number_format = "00000000"
    return cell

number_formatCells( 4, 8 ,  colC) 
'''
-----------------------------------------------------------------------------
                              Main
-----------------------------------------------------------------------------
MOSAIQ Report Generation
MOSAIQ > File > Print Report > Quality Check Lists > QCL Patient Reminder List
Set Date > Tagged > Search "PHY: QA Plan Measurements" > Check Box > Select > Ok
Responsible Location > Tagged > PHY Physics > Group By > QCL Task
QCL Patient Reminder List Report > Print to PDF > Save in folder > Copy pdf file name

'''

   
xlsx_location = r'I:\Physics\aliu3\PatientList.csv'
    
file_path = input('File path: ')

str_path = file_path.replace('\\', '/')

q = p / str_path

sort_files = sorted(q.iterdir(), key=os.path.getmtime, reverse= True)          #sorts by last modifed

latest_file = max([f for f in sort_files], key=lambda item: item.stat().st_ctime)

new_file = 'Newest File: ' +  str(latest_file.absolute())                      #newest file added

list_files = list(sorted(q.glob('**/*.dcm')))                                  #lists .dcm files
      
num_files = len(list_files) - 1                                                #Number of fields

list_stem = str(list_files[0].stem)

ls_int = re.findall(r'(?<=)\d+', list_stem)

ls_str = re.findall(r'(?<=\_|\-)\w+', list_stem)

checkInts = list(filter(lambda x: x.isdigit(), ls_int))

checkStrs = list(filter(lambda x: x.isalpha(), ls_str))
 
if(len(checkInts) == len(ls_int)):
    if 'BeamDose_1' in ls_str:
        field_1 = int(ls_int[1])
        for j, value in enumerate(range(field_1, field_1 + num_files - 1), start= 16):
            ws.cell(row=j, column = colC).value = value          
            ws.cell(row=j, column = colC).number_format = '00'  
    elif any([s.isdigit() for s in ls_str]) == True: 
        field_1 = int(ls_int[-1])    
        for j, value in enumerate(range(field_1, field_1 + num_files - 1), start= 16):
            ws.cell(row=j, column = colC).value = value          
            ws.cell(row=j, column = colC).number_format = '00'                         
    elif 'BeamDose_1' not in ls_str and any([s.isdigit() for s in ls_str]) == False:
        field_1 = int(ls_int[-1])
        for j, value in enumerate(range(field_1, field_1 + num_files - 1), start= 16):
            ws.cell(row=j, column = colC).value = str(value) + str(ls_str[2][-1:])
            ws.cell(row=j, column = colC).number_format = '00' 
    else:
        checkInts == False
        print('Error: Fields')

planName = q.stem                         
ws['C6'].value = planName                                                       #Plan 
Name = q.parts[-2] #Change if wrong
ws['C4'].value = Name
        
if q.parts[5] == 'TB1 Patients':                                              #Machine
    Machine = 'TB1 H193230'
    ws['C8'].value = Machine
elif q.parts[5] == 'TB2 Patients':
    Machine = 'TB2 H193236'
    ws['C8'].value = Machine
elif q.parts[5] == 'TB3 Patients':
    Machine = 'TB3 H193322'
    ws['C8'].value = Machine
elif q.parts[5] == 'TB4 Patients':
    Machine = 'TB4 H193323'
    ws['C8'].value = Machine
else:
    print("Error: Machine")
        
fp_xlsx = '\\' + q.parts[5] + '\\' + Name + '\\' + planName                        
ws['C7'].value = fp_xlsx                                                       #path

last_updated = (time.time() - os.path.getmtime(q) )
last_checked = time.strftime('%H:%M:%S', time.gmtime(last_updated))



df = pd.read_csv(xlsx_location,
                 header = None,
                 names= ['Patient',
                         'MV'
                         ],
                 usecols= [36,53],
                 skip_blank_lines= True,
                 na_filter= False,
                 )

dfDict = df.to_dict(orient='list')

lis = list(map(str.lower, dfDict['Patient']))

lowN = Name.lower()

res = [v for v in lis if v.startswith(lowN)]

newName = res[0].title()[:-10].strip()

newMMRN = re.findall('\d+', res[0])[0]

MV_Dict = dict((zip(dfDict['Patient'], dfDict['MV'])))

MV_ls = [val for key, val in MV_Dict.items() if re.search(Name, key, re.IGNORECASE)] 
   
MV_value = MV_ls[0]

ws['C5'].value = newMMRN                                                       #MMRN
ws['C11'].value = MV_value
ws['C12'].value = MV_value
if any(['6x', '10x', '18x' in MV_value]) == True:
    ws['C13'].value = 600
else:
    ws['C13'].value = 1400

print("Since last update: " + last_checked  ) 
print(df)
print(new_file)

'''
-----------------------------------------------------------------------------
                              Data Validation
-----------------------------------------------------------------------------
ws['XX'] = r'...' are hardcoded - 01/26/23
Conditional formatting in this section
'''

class DataValidation:

    Phantom_Used = DataValidation(type= 'list',
                                  formula1= '"ArcCheck, MapCheck, SRS MapCheck"',
                                  allow_blank= True)
    Phantom_Used.error
    Phantom_Used.errorTitle = 'Invalid Entry'
    Phantom_Used.prompt = 'Choose...'
    ws.add_data_validation(Phantom_Used)
    Phantom_Used.add(ws['C10'])
    
    ws['H10']= r'=IF(C10="MapCheck", 0.05, IF(C10="ArcCheck", 0.1, IF(C10="SRS MapCheck", 0.1, 0.1)))'
    ws['H11']= r'=IF(OR(C10="MapCheck",C10="ArcCheck"), 0.03, IF(C10="SRS MapCheck", 0.02, 0.03))'
    ws['H12']= r'=IF(OR(C10="MapCheck",C10="ArcCheck"), "3 mm", IF(C10="SRS MapCheck", "1 mm", "3 mm"))'
    ws['H13']= 'On'
    
    percent = ws['H10']
    percent.number_format = '0%'
    
    percent1 = ws['H11']
    percent1.number_format = '0%'
    
    DTA_ON = DataValidation(type= "list",
                                formula1= '"On, Off"',
                                allow_blank= False)
    DTA_ON.error
    DTA_ON.errorTitle = 'Invalid Entry'
    ws.add_data_validation(DTA_ON)
    DTA_ON.add(ws['H13'])
    
    fields = [i for i in range(num_files -  1)] 
    
    for i in ws.iter_rows(min_row= 16,
                           max_row= 16 + len(fields) - 1,
                           min_col= colG,
                           max_col= colG
                           ):
        for cell in i:
            cell.value = 'Off'
            
    On_Off = DataValidation(type= 'list',
                                  formula1= '"On, Off"',
                                  allow_blank= False)
    On_Off.error
    On_Off.errorTitle = 'Invalid Entry'
    ws.add_data_validation(On_Off)
    On_Off.add('G16:G24')
    
    #Hardcoded 
    ws['H16'] = r'=IF(AND(ISBLANK(C16),ISBLANK(F16)),"",IF($C$10="SRS MapCheck",IF(AND(ABS(D16)<=5,E16>=90),"Pass","Fail"),IF(G16="Off",IF(AND(ABS(D16)<=3, OR(E16>=90,F16=0)),"Pass","Fail"),IF(AND(AND(ABS(D16)<=5,E16>=90),F16=0),"Pass","Fail"))))'
    ws['H17'] = r'=IF(AND(ISBLANK(C17),ISBLANK(F17)),"",IF($C$10="SRS MapCheck",IF(AND(ABS(D17)<=5,E17>=90),"Pass","Fail"),IF(G17="Off",IF(AND(ABS(D17)<=3, OR(E17>=90,F17=0)),"Pass","Fail"),IF(AND(AND(ABS(D17)<=5,E17>=90),F17=0),"Pass","Fail"))))'
    ws['H18'] = r'=IF(AND(ISBLANK(C18),ISBLANK(F18)),"",IF($C$10="SRS MapCheck",IF(AND(ABS(D18)<=5,E18>=90),"Pass","Fail"),IF(G18="Off",IF(AND(ABS(D18)<=3, OR(E18>=90,F18=0)),"Pass","Fail"),IF(AND(AND(ABS(D18)<=5,E18>=90),F18=0),"Pass","Fail"))))'
    ws['H19'] = r'=IF(AND(ISBLANK(C19),ISBLANK(F19)),"",IF($C$10="SRS MapCheck",IF(AND(ABS(D19)<=5,E19>=90),"Pass","Fail"),IF(G19="Off",IF(AND(ABS(D19)<=3, OR(E19>=90,F19=0)),"Pass","Fail"),IF(AND(AND(ABS(D19)<=5,E19>=90),F19=0),"Pass","Fail"))))'
    ws['H20'] = r'=IF(AND(ISBLANK(C20),ISBLANK(F20)),"",IF($C$10="SRS MapCheck",IF(AND(ABS(D20)<=5,E20>=90),"Pass","Fail"),IF(G20="Off",IF(AND(ABS(D20)<=3, OR(E20>=90,F20=0)),"Pass","Fail"),IF(AND(AND(ABS(D20)<=5,E20>=90),F20=0),"Pass","Fail"))))'
    ws['H21'] = r'=IF(AND(ISBLANK(C21),ISBLANK(F21)),"",IF($C$10="SRS MapCheck",IF(AND(ABS(D21)<=5,E21>=90),"Pass","Fail"),IF(G21="Off",IF(AND(ABS(D21)<=3, OR(E21>=90,F21=0)),"Pass","Fail"),IF(AND(AND(ABS(D21)<=5,E21>=90),F21=0),"Pass","Fail"))))'
    ws['H22'] = r'=IF(AND(ISBLANK(C22),ISBLANK(F22)),"",IF($C$10="SRS MapCheck",IF(AND(ABS(D22)<=5,E22>=90),"Pass","Fail"),IF(G22="Off",IF(AND(ABS(D22)<=3, OR(E22>=90,F22=0)),"Pass","Fail"),IF(AND(AND(ABS(D22)<=5,E22>=90),F22=0),"Pass","Fail"))))'
    ws['H23'] = r'=IF(AND(ISBLANK(C23),ISBLANK(F23)),"",IF($C$10="SRS MapCheck",IF(AND(ABS(D23)<=5,E23>=90),"Pass","Fail"),IF(G23="Off",IF(AND(ABS(D23)<=3, OR(E23>=90,F23=0)),"Pass","Fail"),IF(AND(AND(ABS(D23)<=5,E23>=90),F23=0),"Pass","Fail"))))'
    ws['H24'] = r'=IF(AND(ISBLANK(C24),ISBLANK(F24)),"",IF($C$10="SRS MapCheck",IF(AND(ABS(D24)<=5,E24>=90),"Pass","Fail"),IF(G24="Off",IF(AND(ABS(D24)<=3, OR(E24>=90,F24=0)),"Pass","Fail"),IF(AND(AND(ABS(D24)<=5,E24>=90),F24=0),"Pass","Fail"))))'
    ws['H25'] = r'=IF(AND(ISBLANK(C25),ISBLANK(F25)),"",IF($C$10="SRS MapCheck",IF(AND(ABS(D25)<=5,E25>=90),"Pass","Fail"),IF(G25="Off",IF(AND(ABS(D25)<=3, OR(E25>=90,F25=0)),"Pass","Fail"),IF(AND(AND(ABS(D25)<=5,E25>=90),F25=0),"Pass","Fail"))))'
    
    
    for i in ws.iter_rows(min_row= 16 - 1, #center aligns C15:H25 KEEP HERE 
                           max_row= 16 + len(fields) - 1,
                           min_col= colC,
                           max_col= colH
                           ):
        for cell in i:
            cell.alignment = Alignment(horizontal='center')
            
    redFill = PatternFill(start_color='FFC7CE',
                          end_color='FFC7CE',
                          fill_type='solid')
    
    redFont = Font(bold= True, color= '9c0103')
    
    greenFill = PatternFill(start_color='00CCFFCC',
                          end_color='00CCFFCC',
                          fill_type='solid')
    
    greenFont = Font(bold= True, color= '006100')
    
    ws.conditional_formatting.add('H16:H25',
                                  FormulaRule(formula= ['NOT(ISERROR(SEARCH("Pass",H16)))'],
                                              stopIfTrue= True,
                                              fill= greenFill,
                                              font= greenFont,
                                              )
                                  )    
    
    ws.conditional_formatting.add('H16:H25',
                                  FormulaRule(formula= ['NOT(ISERROR(SEARCH("Fail",H16)))'],
                                              stopIfTrue= True,
                                              fill= redFill,
                                              font= redFont,
                                              )
                                  )

#Save Patient QA worksheet
wb.save(str(q) + '/' + 'Patient QA ' + newName + '.xlsx')

wb.close()

#End
