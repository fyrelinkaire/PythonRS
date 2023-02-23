
"""
Created on Tue Sep 27 09:06:32 2022
Uploaded on Tue Oct 18 13:50:00 2022 to RS
IMPORTANT: WORKSHEET 2 DOES NOT EXIST BUT LEFT IN FOR REFERENCE
@author: ALiu3
"""

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import colors
from openpyxl.styles import Font, Alignment, PatternFill #, Color, Border
#from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import FormulaRule #, ColorScaleRule, CellIsRule, Rule
from pathlib import Path
import os
import time


'''
-----------------------------------------------------------------------------
                            Initialization
-----------------------------------------------------------------------------
'''
p = Path('.')                            #Pathlib module
wb = Workbook()                          #Open Patient QA worksheet
ws = wb.active                           #Set active worksheet
# ws2 = wb.create_sheet('Sheet2')        #NOT USED

'''
-----------------------------------------------------------------------------
                            Sheet 1 Formatting
-----------------------------------------------------------------------------
'''
rs4 = 4   
rs10 = 10  
rs16 = 16  
rs27 = 27
rs32 = 32
re8 = 8   
re13 = 13  
re25 = 25
re29 = 29
re34 = 34
col1 = 1  #A
col2 = 2  #B
col3 = 3  #C
col4 = 4  #D
col5 = 5  #E
col6 = 6  #f
col7 = 7  #G
col8 = 8  #H
col9 = 9  #I
ws['A4'] = 'Patient Name'
ws['A5'] = 'Patient URN'
ws['A6'] = 'Plan Name'
ws['A7'] = 'Plan folder'
ws['A8'] = 'Machine'
ws['A10'] = 'Phantom Used'
ws['A11'] = 'Dose Calibration'
ws['A12'] = 'Array Calibration'
ws['A13'] = 'Dose Rate (MU/min)'
ws['G10'] = 'Threshold %'
ws['G11'] = '% Dose Difference'
ws['G12'] = 'Dist to Agreement'
ws['G13'] = '3D DTA Mode'
ws['C15'] = 'Field'
ws['D15'] = 'Dose' + ' ' u'\u2206' + '%'
ws['E15'] = 'Passing %'
ws['F15'] = '#pts y>2'
ws['G15'] = 'Global'
ws['H15'] = 'P/F'
ws['A27'] = '(1) Absolute point dose difference is ≤ 3%'
ws['A28'] = '(2) Absolute point dose difference is ≤ 5%'
ws['A29'] = '(3) % points passing the absolute y test is ≥ 90%'
ws['A30'] = '(4) No points with y > 2 on the absolute y test'
ws['A32'] = 'Insert comments here…'
ws['I27'] = 'Passes if (2) & (3) or (4) are satisfied if global off'
ws['I28'] = 'Passes if (1) & (3) & (4) are satisfied if global on'
ws['I29'] = 'Otherwise fails'

title = ws['A2'] # title 
title.font = Font(sz=18)
title.alignment = Alignment(horizontal='center')
ws.merge_cells('A2:I2')
ws['A2'] = 'Patient QA Worksheet'
   
for rows in ws.iter_rows(min_row= rs4, #color cells C4:C8 
                         max_row= re8,
                         min_col= col2,
                         max_col= col8):
  for cell in rows:
    cell.fill = PatternFill(patternType='solid', fgColor=colors.Color(rgb='EFF7FF'))
    
for rows in ws.iter_rows(min_row= rs10, #color cells H10:H13
                         max_row= re13,
                         min_col= col8,
                         max_col= col8):
  for cell in rows:
    cell.fill = PatternFill(patternType='solid', fgColor=colors.Color(rgb='EFF7FF'))
    
for rows in ws.iter_rows(min_row= rs10, #color cells C10:C13
                         max_row= re13,
                         min_col= col3,
                         max_col= col4):
  for cell in rows:
    cell.fill = PatternFill(patternType='solid', fgColor=colors.Color(rgb='EFF7FF'))
    
for rows in ws.iter_rows(min_row= rs16, #color cells C16:C25
                         max_row= re25,
                         min_col= col3,
                         max_col= col8):
  for cell in rows:
    cell.fill = PatternFill(patternType='solid', fgColor=colors.Color(rgb='EFF7FF'))
    
for rows in ws.iter_rows(min_row= rs32, #color cells A32:A34
                         max_row= re34,
                         min_col= col1,
                         max_col= col9):
  for cell in rows:
    cell.fill = PatternFill(patternType='solid', fgColor=colors.Color(rgb='EFF7FF'))    
    

for i in range(rs4, re8 + 1): #replaces A4:A8, merges col1 + col2
    ws.merge_cells(start_row = i,
                    start_column = col1,
                    end_row = i,
                    end_column = col2,
        )
for i in range(rs10, re13 + 1): #replaces A10:A13, merges col1 + col2
    ws.merge_cells(start_row = i,
                    start_column = col1,
                    end_row = i,
                    end_column = col2,
        )    
for i in range(rs4, re8 + 1): #replaces C4:C8, merges col4 + col8
    ws.merge_cells(start_row = i,
                    start_column = col3,
                    end_row = i,
                    end_column = col8,
        )   
for i in range(rs10, re13 + 1): #replaces C10:C13, merges col4 + col8
    ws.merge_cells(start_row = i,
                    start_column = col3,
                    end_row = i,
                    end_column = col4,
        )  

for i in range(rs4, re8 + 1): #right aligns A4:A8
    align = ws.cell(row = i, column = col1)
    align.alignment = Alignment(horizontal='right')
    i += 1

for i in range(rs10, re13 + 1): #right aligns A10:A13
    align = ws.cell(row = i, column = col1)
    align.alignment = Alignment(horizontal='right')
    i += 1

for i in range(rs10, re13 + 1): #right aligns G10:G13
    align = ws.cell(row = i, column = col7)
    align.alignment = Alignment(horizontal='right')
    i += 1

for i in range(rs10, re13 + 1): #center aligns C10:C13
    align = ws.cell(row = i, column = col3)
    align.alignment = Alignment(horizontal='center')
    i += 1

for i in range(rs27, re29 + 1): #left aligns I27:I29
    align = ws.cell(row = i, column = col9)
    align.alignment = Alignment(horizontal='right')
    i += 1
    
for i in range(rs10, re13 + 1): #center aligns H10:H13
    align = ws.cell(row = i, column = col8)
    align.alignment = Alignment(horizontal='center')
    i += 1

'''
-----------------------------------------------------------------------------
Sheet 2 Formatting Depreciated
-----------------------------------------------------------------------------
'''
# ws2['A1'] = 'Machines'
# ws2['A2'] = 'Choose...'
# ws2['A3'] = 'TB1_H193230'
# ws2['A4'] = 'TB2_H193236'
# ws2['A5'] = 'TB3_H193322' 
# ws2['A6'] = 'TB4_H193323'
# ws2['B1'] = 'Phantoms'
# ws2['B2'] = 'Choose...'
# ws2['B3'] = 'ArcCheck'
# ws2['B4'] = 'Mapcheck'
# ws2['B5'] = 'SRS MapCheck'
# ws2['C1'] = 'PassFaii'
# ws2['C2'] = 'Pass'
# ws2['C3'] = 'Fail'
# ws2['D1'] = 'YesNo'
# ws2['D2'] = 'Yes'
# ws2['D3'] = 'No'
# ws2['E1'] = 'OnOff'
# ws2['E2'] = 'On'
# ws2['E3'] = 'Off'
# ws2['F1'] = 'Thresholds'
# ws2['F2'] = 'Choose...'
# ws2['F3'] = '5%'
# ws2['F4'] = '10%'
# ws2['F5'] = '15%'
# ws2['F6'] = '20%'
# ws2['G1'] = 'DoseDiffs'
# ws2['G2'] = 'Choose...'
# ws2['G3'] = '1%'
# ws2['G4'] = '2%'
# ws2['G5'] = '3%'
# ws2['G6'] = '4%'
# ws2['G7'] = '5%'
'''
-----------------------------------------------------------------------------
                              Function
-----------------------------------------------------------------------------
fields cells are formatted here
'''
file_path = input('File path: ')
str_path = file_path.replace('\\', '/')
q = p / str_path

sort_files = sorted(q.iterdir(), key=os.path.getmtime, reverse= True) #sorts by last modifed
# file_list = list(q.glob('**/*.DCM')) #default list

latest_file = max([f for f in sort_files], key=lambda item: item.stat().st_ctime)

new_file = 'Newest file: ' +  str(latest_file.absolute())
  
    
if q.name == q.stem:
    
    num_files = len(list(q.glob('**/*.DCM'))) - 1 #Remove RPM file

    fields = [i for i in range(num_files -  1)]

    for j, value in enumerate(fields, start=rs16):
        ws.cell(row=j, column = col3).value = value + 1  
        ws.cell(row=j, column = col3).number_format = '00'
        
    
        list_files = list(sorted(q.iterdir(), key=os.path.getmtime))
        
    ws['C4'].value = ' ' + list_files[1].parts[6] #Name
    ws['C5'].value = ' ' + list_files[1].stem[0:8] #URN
    ws['C6'].value = ' ' + q.name  #Plan Name
    ws['C7'].value = ' ' + '\\' + list_files[1].parts[5] + '\\' + list_files[1].parts[6] + '\\' + str(q.name)  #path
    ws['C8'].value = ' ' + list_files[1].parts[5] # Machine

file_list = list(q.glob('**/*.DCM')) #default list
last_updated = (time.time() - os.path.getmtime(q) )
last_checked = time.strftime('%H:%M:%S', time.gmtime(last_updated))

for f in file_list: #Dose files
    if f.stat().st_mtime > last_updated:
        print(f.absolute())            
    else:
        print('No Updates in Dose Files') 
# for item in q.glob('**/*'): #All file types
#     if item.is_file():
#         print (str(item.absolute()))
#     else:
#         print('No Updates')
print("Since last update: " + last_checked  ) 

print(new_file)

'''
-----------------------------------------------------------------------------
                              Data Validation
-----------------------------------------------------------------------------
ws['XX'] = r'...' are hardcoded - 01/26/23
Conditional formatting in this section
'''
Phantom_Used = DataValidation(type= 'list',
                              formula1= '"ArcCheck, MapCheck, SRS MapCheck"',
                              allow_blank= True)
Phantom_Used.error
Phantom_Used.errorTitle = 'Invalid Entry'
Phantom_Used.prompt = 'Choose...'
ws.add_data_validation(Phantom_Used)
Phantom_Used.add(ws['C10'])

ws['H10']= r'=IF(C10="MapCheck", 0.05, IF(C10="ArcCheck", 0.1, IF(C10="SRS MapCheck", 0.1, 0.1)))'
ws['H11']= r'=IF(OR(C10="MapCheck",C10="ArcCheck"), 0.03, IF(C10="SRS MapCheck", 0.02, 0.03))'
ws['H12']= r'=IF(OR(C10="MapCheck",C10="ArcCheck"), "3 mm", IF(C10="SRS MapCheck", "1 mm", "3 mm"))'
ws['H13']= 'On'

percent = ws['H10']
percent.number_format = '0%'

percent1 = ws['H11']
percent1.number_format = '0%'

DTA_ON = DataValidation(type= "list",
                            formula1= '"On, Off"',
                            allow_blank= False)
DTA_ON.error
DTA_ON.errorTitle = 'Invalid Entry'
ws.add_data_validation(DTA_ON)
DTA_ON.add(ws['H13'])

for i in ws.iter_rows(min_row= rs16,
                       max_row= rs16 + len(fields) - 1,
                       min_col= col7,
                       max_col= col7
                       ):
    for cell in i:
        cell.value = 'Off'
        
On_Off = DataValidation(type= 'list',
                              formula1= '"On, Off"',
                              allow_blank= False)
On_Off.error
On_Off.errorTitle = 'Invalid Entry'
ws.add_data_validation(On_Off)
On_Off.add('G16:G24')

#Hardcoded 
ws['H16'] = r'=IF(AND(ISBLANK(C16),ISBLANK(F16)),"",IF($C$10="SRS MapCheck",IF(AND(ABS(D16)<=5,E16>=90),"Pass","Fail"),IF(G16="Off",IF(AND(ABS(D16)<=3, OR(E16>=90,F16=0)),"Pass","Fail"),IF(AND(AND(ABS(D16)<=5,E16>=90),F16=0),"Pass","Fail"))))'
ws['H17'] = r'=IF(AND(ISBLANK(C17),ISBLANK(F17)),"",IF($C$10="SRS MapCheck",IF(AND(ABS(D17)<=5,E17>=90),"Pass","Fail"),IF(G17="Off",IF(AND(ABS(D17)<=3, OR(E17>=90,F17=0)),"Pass","Fail"),IF(AND(AND(ABS(D17)<=5,E17>=90),F17=0),"Pass","Fail"))))'
ws['H18'] = r'=IF(AND(ISBLANK(C18),ISBLANK(F18)),"",IF($C$10="SRS MapCheck",IF(AND(ABS(D18)<=5,E18>=90),"Pass","Fail"),IF(G18="Off",IF(AND(ABS(D18)<=3, OR(E18>=90,F18=0)),"Pass","Fail"),IF(AND(AND(ABS(D18)<=5,E18>=90),F18=0),"Pass","Fail"))))'
ws['H19'] = r'=IF(AND(ISBLANK(C19),ISBLANK(F19)),"",IF($C$10="SRS MapCheck",IF(AND(ABS(D19)<=5,E19>=90),"Pass","Fail"),IF(G19="Off",IF(AND(ABS(D19)<=3, OR(E19>=90,F19=0)),"Pass","Fail"),IF(AND(AND(ABS(D19)<=5,E19>=90),F19=0),"Pass","Fail"))))'
ws['H20'] = r'=IF(AND(ISBLANK(C20),ISBLANK(F20)),"",IF($C$10="SRS MapCheck",IF(AND(ABS(D20)<=5,E20>=90),"Pass","Fail"),IF(G20="Off",IF(AND(ABS(D20)<=3, OR(E20>=90,F20=0)),"Pass","Fail"),IF(AND(AND(ABS(D20)<=5,E20>=90),F20=0),"Pass","Fail"))))'
ws['H21'] = r'=IF(AND(ISBLANK(C21),ISBLANK(F21)),"",IF($C$10="SRS MapCheck",IF(AND(ABS(D21)<=5,E21>=90),"Pass","Fail"),IF(G21="Off",IF(AND(ABS(D21)<=3, OR(E21>=90,F21=0)),"Pass","Fail"),IF(AND(AND(ABS(D21)<=5,E21>=90),F21=0),"Pass","Fail"))))'
ws['H22'] = r'=IF(AND(ISBLANK(C22),ISBLANK(F22)),"",IF($C$10="SRS MapCheck",IF(AND(ABS(D22)<=5,E22>=90),"Pass","Fail"),IF(G22="Off",IF(AND(ABS(D22)<=3, OR(E22>=90,F22=0)),"Pass","Fail"),IF(AND(AND(ABS(D22)<=5,E22>=90),F22=0),"Pass","Fail"))))'
ws['H23'] = r'=IF(AND(ISBLANK(C23),ISBLANK(F23)),"",IF($C$10="SRS MapCheck",IF(AND(ABS(D23)<=5,E23>=90),"Pass","Fail"),IF(G23="Off",IF(AND(ABS(D23)<=3, OR(E23>=90,F23=0)),"Pass","Fail"),IF(AND(AND(ABS(D23)<=5,E23>=90),F23=0),"Pass","Fail"))))'
ws['H24'] = r'=IF(AND(ISBLANK(C24),ISBLANK(F24)),"",IF($C$10="SRS MapCheck",IF(AND(ABS(D24)<=5,E24>=90),"Pass","Fail"),IF(G24="Off",IF(AND(ABS(D24)<=3, OR(E24>=90,F24=0)),"Pass","Fail"),IF(AND(AND(ABS(D24)<=5,E24>=90),F24=0),"Pass","Fail"))))'
ws['H25'] = r'=IF(AND(ISBLANK(C25),ISBLANK(F25)),"",IF($C$10="SRS MapCheck",IF(AND(ABS(D25)<=5,E25>=90),"Pass","Fail"),IF(G25="Off",IF(AND(ABS(D25)<=3, OR(E25>=90,F25=0)),"Pass","Fail"),IF(AND(AND(ABS(D25)<=5,E25>=90),F25=0),"Pass","Fail"))))'


for i in ws.iter_rows(min_row= rs16 - 1, #center aligns C15:H25 KEEP HERE 
                       max_row= rs16 + len(fields) - 1,
                       min_col= col3,
                       max_col= col8
                       ):
    for cell in i:
        cell.alignment = Alignment(horizontal='center')
        
redFill = PatternFill(start_color='FFC7CE',
                      end_color='FFC7CE',
                      fill_type='solid')

redFont = Font(bold= True, color= '9c0103')

greenFill = PatternFill(start_color='00CCFFCC',
                      end_color='00CCFFCC',
                      fill_type='solid')

greenFont = Font(bold= True, color= '006100')

ws.conditional_formatting.add('H16:H25',
                              FormulaRule(formula= ['NOT(ISERROR(SEARCH("Pass",H16)))'],
                                          stopIfTrue= True,
                                          fill= greenFill,
                                          font= greenFont,
                                          )
                              )    

ws.conditional_formatting.add('H16:H25',
                              FormulaRule(formula= ['NOT(ISERROR(SEARCH("Fail",H16)))'],
                                          stopIfTrue= True,
                                          fill= redFill,
                                          font= redFont,
                                          )
                              )

#Save Patient QA worksheet
wb.save(str(q) + '/' + "Patient QA Worksheet.xlsx")

wb.close()

#End