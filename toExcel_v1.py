
"""
Created on Tue Sep 27 09:06:32 2022
Uploaded on Tue Oct 18 13:50:00 2022
@author: ALiu3
"""

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path
from connect import *

'''
run_console > statetree.RunStateTree() > Copy to Clipboard
'''

Patient_Name = get_current('Patient')
Plan_Name = get_current('Plan')
Machine_Name = patient.Cases[0].TreatmentPlans[2].BeamSets[0].MachineReference.MachineName
TB = Machine_Name[:3]

#Returns TB, Patient Name, Plan Name, URN
#Move to working directory - patient folder
p = Path('.')
#q = [x for x in p.iterdir() if x.is_dir()] shows directory files
z = p / TB / Patient_Name / Plan_Name 


def Upload():
        
            
    #Open Patient QA worksheet
    wb = load_workbook(filename = "PatientQA_Worksheet.xlsx")
    
    ws = wb.active
    
    #Find URN from folder path

    for subdir in sorted(Path(z).iterdir()):
        f = str(subdir)
        g = str(z)
        if g in f:
            strReplace = f.replace(g,'')
        URNstr = strReplace[1:9]
        if sum(c.isdigit() for c in URNstr) == 8:
            URN = URNstr
        elif sum(c.isdigit() for c in URNstr) != 8:
            strSubdirpart = str(subdir.parts[3])
            f.replace(strSubdirpart, URNstr)

    #Set values in spreasheet
    ws['C4'].value = z.parts[1] #Patient Name
    ws['C5'].value = URN        #URN
    ws['C6'].value = z.parts[2] #Plan Name
    ws['C7'].value = g          #Plan Folder
    ws['C8'].value = z.parts[0] #Machine
    
    #DataValidation for Phantom
    dv = DataValidation(type="list", formula1='"ArcCheck, MapCheck, SRS MapCheck"', 
                            allow_blank = True)
    #Error + Prompt Msgs
    dv.error = "Check yo spelling"
    dv.errorTitle = 'Entry Invalid'
    dv.prompt = "Select from list"
    dv.promptTitle = "Choose"
    ws.add_data_validation(dv) #needs wb.active

    #DV to Wkbk
    dv.add(ws['C10'])
    
    #add number of fields
    num_files = len(list(z.glob('*')))

    fields = [i for i in range(num_files-1)]

    row16 = 16
    columnC = 3

    for j, value in enumerate(fields, start=row16):
        ws.cell(row=j, column = columnC).value = value + 1  
    
    #Save Patient QA worksheet
    wb.save(filename = "PatientQA_Worksheet.xlsx")
    wb.close()
    
    return 


